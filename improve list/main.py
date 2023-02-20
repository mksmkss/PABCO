import datetime
import json
import os
import sys

import openpyxl as px
import pandas as pd
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.styles.alignment import Alignment

main_path = os.path.dirname(sys.argv[0])
print(f"This program is working on {main_path}")


excels_list = []

new_excel = {}


def setting_files():
    global inventory_supply_df, inventory_ahead_df, permanent_df, permanent_db_df, subject_path, output_path
    _settings = open(f"{main_path}/settings.json", "r", encoding="utf-8")
    # json -> 辞書型
    settings = json.load(_settings)

    # Excelファイルの読み込みのための設定
    inventory_path = settings["inventory_excel"]
    inventory_excel = pd.read_excel(inventory_path, sheet_name=None)
    inventory_supply_df = inventory_excel["支給品在庫検索"].drop(
        inventory_excel["支給品在庫検索"].index[0:3]
    )
    inventory_ahead_df = inventory_excel["先行部品検索"].drop(
        inventory_excel["先行部品検索"].index[0:3]
    )
    permanent_path = settings["permanent_excel"]
    permanent_excel = pd.read_excel(permanent_path, sheet_name=None)
    permanent_df = permanent_excel["常置品入出庫表 (新)"].drop(
        permanent_excel["常置品入出庫表 (新)"].index[0:3]
    )
    permanent_db = settings["permanent_db"]
    permanent_db_excel = pd.read_excel(permanent_db, sheet_name=None)
    permanent_db_df = permanent_db_excel["DB"].drop(permanent_db_excel["DB"].index[0:3])

    # フォルダのパスを取得
    subject_path = settings["subject_folder"]
    output_path = settings["output_folder"]


def get_excel_list():
    """
    ディレクトリ内のExcelファイルをリスト型で取得する関数
    """
    # os.listdir() -> ディレクトリ内のファイル名をリストで返す
    for file in os.listdir(subject_path):
        if file.endswith(".xlsx") or file.endswith(".xls") or file.endswith(".xlsm"):
            excels_list.append(file)
    return excels_list


def get_products_list(dfs, sheet):
    """
    物品のデータをリスト型で取得する関数
    dfs:Excelファイルのデータフレーム
    sheet:シート名
    """
    products = dfs[sheet].values.tolist()
    del products[:6]
    return products


def permanent_edit(supp, pro_num, dfs, sheet, pro_row):
    pro_num_list = permanent_db_df.iloc[:, 0].tolist()
    # K1送付済みかどうか
    if pro_num in pro_num_list:
        isDone = bool(
            type(permanent_db_df.iloc[pro_num_list.index(pro_num), 4]) != float
        )
    else:
        isDone = False
    if supp == "発注点":
        dfs[sheet].iat[pro_row + 6, 10] = "ﾊﾟﾌﾞｺｰｷから納入後、発送願います"
        # 空欄かつdbにあったら
    elif type(supp) == float and pro_num in pro_num_list:
        dfs[sheet].iat[pro_row + 6, 10] = permanent_db_df.iloc[
            pro_num_list.index(pro_num), 5
        ]
        # K1送付済み
    elif isDone == True:
        dfs[sheet].iat[pro_row + 6, 10] = "K1在庫 発送不要"
    elif supp == "常置品":
        if pro_num in pro_num_list:
            isTent = bool(permanent_db_df.iloc[pro_num_list.index(pro_num), 3] > 0)
            if isTent == True:
                dfs[sheet].iat[pro_row + 6, 10] = "ﾃﾝﾄ在庫 発送お願いします"
            else:
                # 発注点管理かそれ以外
                if permanent_db_df.iloc[pro_num_list.index(pro_num), 1] == "発注点管理":
                    dfs[sheet].iat[pro_row + 6, 10] = "ﾊﾟﾌﾞｺｰｷから納入後、発送お願いします"
                    dfs[sheet].iat[pro_row + 6, 2] = "発注点"
                else:
                    dfs[sheet].iat[pro_row + 6, 2] = "個別発注"
        else:
            dfs[sheet].iat[pro_row + 6, 10] = "常置品DBに登録されていません.確認お願いします"


def search_place(arr, supp, pro_num, dfs, sheet, pro_row):
    """
    場所を検索し，セルに書き込むまでの関数
    arr:補足, supp:検索する補足
    """
    row = 0
    if arr == "支給":
        # suppの一文字目を削除
        supp = supp[1:]
        # 支給品在庫検索シートの支給品からエリア別を検索
        col_supply_list = inventory_supply_df.iloc[:, 2].tolist()
        for i in col_supply_list:
            if i == supp:
                # rowはinventory_supply_df内の行番号
                row = col_supply_list.index(i)
                place = inventory_supply_df.iloc[row, 0]
                dfs[sheet].iat[pro_row + 6, 15] = place
                break
    elif arr == "先行":
        # 先行部品検索シートの先行部品からエリア別を検索
        col_ahead_list = inventory_ahead_df.iloc[:, 3].tolist()
        for i in col_ahead_list:
            if i == supp:
                # numはinventory_ahead_df内の行番号
                num = col_ahead_list.index(i)
                place = f"{inventory_ahead_df.iloc[num, 0]} {inventory_ahead_df.iloc[num, 1]}"
                dfs[sheet].iat[pro_row + 6, 15] = place
                break
    elif arr == "部品":
        # 常置品在庫検索シートの常置品からエリア別を検索
        col_permanent_list = permanent_df.iloc[:, 1].tolist()
        for i in col_permanent_list:
            if i == pro_num:
                # numはpermanent_df内の行番号
                num = col_permanent_list.index(i)
                place = permanent_df.iloc[num, 0]
                dfs[sheet].iat[pro_row + 6, 15] = place
                break


def add_job_figure(dfs, sheet):
    """
    作業番号を追加する関数
    """
    # 作業番号の取得
    _job_figure = dfs[sheet].iat[1, 2]
    job_figure = _job_figure[3:8]
    # 作業図番号の追加
    dfs[sheet].iat[2, 15] = job_figure


def modify_date(product, dfs, sheet, row):
    """
    日付の形式を修正する関数
    2023-07-15 00:00:00->7/15
    """
    _date = str(product[7])
    # 日付の形式を修正
    _date = _date.replace("-", "/")
    date = _date[5:10]
    parts = date.split("/")
    if len(parts[0]) == 0:
        pass
    else:
        if parts[0].startswith("0"):
            parts[0] = parts[0].replace("0", "")
        if parts[1].startswith("0"):
            parts[1] = parts[1].replace("0", "")
    date = "/".join(parts)
    dfs[sheet].iat[row + 6, 7] = date


def main():
    # 補足を保存
    pre_supp = ""
    setting_files()

    for excel in get_excel_list():
        # Excelファイルを読み込む. dfsには辞書型でシート名とデータフレームが入る
        dfs = pd.read_excel(f"{subject_path}/{excel}", sheet_name=None)
        # ファイル名と拡張子を分ける
        excel_name, excel_ext = os.path.splitext(excel)
        for sheet in dfs.keys():
            print(f"{sheet}")
            # まず列を追加
            date = str(datetime.date.today()).replace("-", "/")
            dfs[sheet][date] = ""
            # 一部項目追加（変更）
            dfs[sheet].iat[5, 8] = "数量\n/１台"
            dfs[sheet].iat[5, 9] = "合計\n数量"
            dfs[sheet].iat[5, 15] = "場所"
            products = get_products_list(dfs, sheet)
            # 一番最初の補足を取得
            pre_supp = products[0][1]
            row = 0
            add_job_figure(dfs, sheet)
            # K1工事かどうか
            isK1 = bool(dfs[sheet].iat[1, 5][-4:] == "Ｋ１工事")
            for product in products:
                """
                product[0]:No.
                product[1]:手配
                product[2]:補足
                product[3]:部品番号
                """
                # リストの中での行番号を取得
                row = products.index(product)
                # 日付を整形
                modify_date(product, dfs, sheet, row)
                # nanはfloat64型なのでnanの判定はこうする．i=="nan"ではだめ．
                if (type(product[2]) != float) and (product[2] != pre_supp):
                    # 場所を検索してセルに書き込む
                    search_place(product[1], product[2], product[3], dfs, sheet, row)
                    # 前の補足を更新
                    pre_supp = product[2]
                # 　次工程に追加
                if isK1 == True and product[1] == "部品":
                    permanent_edit(product[2], product[3], dfs, sheet, row)

        # Excelファイルを書き込む
        new_path = f"{output_path}/{excel_name}_output.xlsx"
        with pd.ExcelWriter(new_path) as writer:
            for sheet in dfs.keys():
                dfs[sheet].to_excel(writer, sheet_name=sheet, index=False)

        """
        ~~整形~~
        """
        # 定義
        center_alignment = Alignment(horizontal="center", vertical="center")
        left_alignment = Alignment(horizontal="left", vertical="center")
        fill = PatternFill(patternType="solid", fgColor="d8d8d8")
        wrap_text = Alignment(wrap_text=True)
        side = Side(style="thin", color="000000")
        thick = Side(style="thick", color="000000")
        border = Border(top=side, bottom=side, left=side, right=side)
        border_none = Border(top=None, bottom=None, left=None, right=None)
        thick_cell = Border(top=thick, bottom=side, left=side, right=side)

        wb = px.load_workbook(new_path)

        merge_cell_list = [
            "A1:O1",
            "A2:P2",
            "A3:B3",
            "C3:D3",
            "F3:M3",
            "O3:P3",
            "A4:B4",
            "C4:D4",
            "F4:O4",
            "A5:B5",
            "C5:D5",
            "J5:P5",
            "H5:I5",
            "A6:P6",
            "E7:F7",
        ]
        col_list = [
            "A",
            "B",
            "C",
            "D",
            "E",
            "F",
            "G",
            "H",
            "I",
            "J",
            "K",
            "L",
            "M",
            "N",
            "O",
            "P",
        ]
        gray_cell_list = [
            "A3",
            "E3",
            "N3",
            "A4",
            "E4",
            "A5",
            "E5",
            "G5",
            "A7",
            "B7",
            "C7",
            "D7",
            "E7",
            "F7",
            "G7",
            "H7",
            "I7",
            "J7",
            "K7",
            "L7",
            "M7",
            "N7",
            "O7",
            "P7",
        ]
        height_tuple = [(1, 30), (2, 11), (3, 30), (4, 62), (5, 30), (6, 11), (7, 39)]
        width_tuple = [
            ("A", 4.3),
            ("B", 7.7),
            ("C", 9.7),
            ("D", 17.7),
            ("E", 11.7),
            ("F", 15.7),
            ("G", 17.7),
            ("H", 8.0),
            ("I", 7.7),
            ("J", 7.7),
            ("K", 8.7),
            ("L", 6.7),
            ("M", 8.7),
            ("N", 11.7),
            ("O", 15.7),
            ("P", 18.7),
        ]
        wrap_cell_list = ["I7", "J7"]
        without_border_list = [
            "A1",
            "B1",
            "C1",
            "D1",
            "E1",
            "F1",
            "G1",
            "H1",
            "I1",
            "J1",
            "K1",
            "L1",
            "M1",
            "N1",
            "O1",
            "P1",
            "A2",
            "B2",
            "C2",
            "D2",
            "E2",
            "F2",
            "G2",
            "H2",
            "I2",
            "J2",
            "K2",
            "L2",
            "M2",
            "N2",
            "O2",
            "P2",
            "J5",
            "K5",
            "L5",
            "M5",
            "N5",
            "O5",
            "P5",
            "A6",
            "B6",
            "C6",
            "D6",
            "E6",
            "F6",
            "G6",
            "H6",
            "I6",
            "J6",
            "K6",
            "L6",
            "M6",
            "N6",
            "O6",
            "P6",
        ]

        for sheet in wb.sheetnames:
            # すべてのセルのフォントサイズを変える
            for row in wb[sheet].rows:
                for cell in row:
                    cell.font = Font(size=18)
            # A1のセル
            wb[sheet].cell(1, 1).alignment = left_alignment
            wb[sheet].cell(1, 1).font = Font(bold=True, size=24)
            # P4のセル
            wb[sheet].cell(4, 16).font = Font(bold=True, size=24)
            # セルの中央揃え
            for col in col_list:
                for row in [3, 4, 5, 7]:
                    wb[sheet][f"{col}{row}"].alignment = center_alignment
            # セルの中央揃え(H8~)
            for row in wb[sheet].iter_rows(min_row=8):
                for cell in row:
                    cell.alignment = center_alignment
            """ 
            次工程を左揃えに
            wb[sheet].iter_rows(min_row=8)がgeneratorなのでlen()で長さを取得できない->sumをつかう
            """
            len = sum(1 for _ in wb[sheet].iter_rows(min_row=8))
            for row in range(8, len + 8):
                wb[sheet][f"K{row}"].alignment = left_alignment
            # セルの結合
            for cell in merge_cell_list:
                wb[sheet].merge_cells(cell)
            # セルの結合(E7:F7,E8:F8,E9:F9,...)
            for row in range(7, 58):
                wb[sheet].merge_cells(f"E{row}:F{row}")
            # セルの色を変える&中央揃え&フォントサイズを一部14に
            for cell in gray_cell_list:
                wb[sheet][cell].fill = fill
                wb[sheet][cell].alignment = center_alignment
                wb[sheet][cell].font = Font(size=14)
            # セルの高さを変える
            for height in height_tuple:
                wb[sheet].row_dimensions[height[0]].height = height[1]
            for width in width_tuple:
                wb[sheet].column_dimensions[width[0]].width = width[1]
            # セルの折り返し
            for cell in wrap_cell_list:
                wb[sheet][cell].alignment = wrap_text
            # セルの罫線
            for row in wb[sheet].rows:
                for cell in row:
                    # cell.coordinate = A1, cell.value = そのセルの中身
                    if cell.coordinate in without_border_list:
                        wb[sheet][cell.coordinate].border = border_none
                    else:
                        cell.border = border
            # セルの太線
            pre_value = ""
            for row in wb[sheet].rows:
                if row[1].value == "支給" and pre_value == "部品":
                    for cell in row:
                        cell.border = thick_cell
                pre_value = row[1].value
        wb.save(new_path)


if __name__ == "__main__":
    main()
